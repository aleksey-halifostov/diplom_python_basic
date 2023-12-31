import openpyxl
import json
import secondary_functions as function


# Главная исполняемая функция
# выполняет роль главного меню
# запрашивает у пользователя действие и вызывает соответствующие функции
def main():
    func_to_call = {1: new_person, 2: find_in_data, 3: from_other_to_data, 4: from_data_to_other, 5: to_json}
    while True:
        print("Выберите действие (введите номер)")
        print("1. Ввести новую запись")
        print("2. Поиск в БД")
        print("3. Загрузить данные в БД из файла")
        print("4. Сохранить данные из БД в файл")
        print("5. Экспортировать данные в json формат")
        print('-' * 50)
        print("0. Выход")
        choice = input("Ваш выбор: ")

        if not choice.isdigit() or int(choice) > 5:
            print("Не корректный ввод")
            continue
        elif int(choice) == 0:
            break

        func_to_call[int(choice)]()


# Декоратор открывает и закрывает базу данных после работы с ней
def open_and_closing_data(func):
    def dec_func():

        book = openpyxl.load_workbook("data.xlsx")
        sheet = book["Sheet"]
        try:
            func(sheet)
        finally:
            book.save("data.xlsx")
            book.close()

    return dec_func


# Записывает информацию о новом человеке в базу данных
@open_and_closing_data
def new_person(sheet):
    # Создание объекта класса Person, подготовка к записи
    new_obj = function.get_data_from_operator()
    row = function.find_empty_row(sheet)

    # Запись новых данных, сохранение и закрытие файла
    for ind, elem in enumerate(new_obj.__dict__.values()):
        sheet.cell(row, ind + 1).value = elem


# Выполняет поиск в базе данных по введенному оператором запросу
# печатает все результаты, которые удовлетворяют поиск
@open_and_closing_data
def find_in_data(sheet):
    to_find = input("Введите ФИО: ").upper()
    print("-" * 50)
    male_or_female = {"мужчина": ("Родился:", "Умер:"), "женщина": ("Родилась:", "Умерла:")}

    for row in range(1, sheet.max_row + 1):
        if to_find in str(sheet.cell(row, 1).value).upper():
            death_str = ""
            birthday = function.date(sheet.cell(row, 3).value)
            last_day = function.date(sheet.cell(row, 4).value)
            if sheet.cell(row, 4).value:
                death_str = f" {male_or_female[sheet.cell(row, 2).value][1]} {last_day}"
            print(
                f"{sheet.cell(row, 1).value} {function.def_age(birthday, last_day)}, {sheet.cell(row, 2).value}. " +
                f"{male_or_female[sheet.cell(row, 2).value][0]} {birthday}"
                + death_str)
    print("Показаны все удовлетворяющие поиск результаты, или результаты отсутствуют...")
    print("-" * 50)


# Функция загружает данные из БД в файл other.xlsx
@open_and_closing_data
def from_data_to_other(sheet):
    other_book = openpyxl.load_workbook("other.xlsx")
    other_sheet = other_book["Sheet"]
    row_other = function.find_empty_row(other_sheet)

    try:
        for row in range(1, sheet.max_row + 1):
            for column in range(1, 5):
                other_sheet.cell(row_other, column).value = sheet.cell(row, column).value

            row_other += 1

    finally:
        other_book.save("other.xlsx")
        other_book.close()


# Функция загружает данные из файла other.xlsx в БД
@open_and_closing_data
def from_other_to_data(sheet):
    other_book = openpyxl.load_workbook("other.xlsx")
    other_sheet = other_book["Sheet"]
    row_other = function.find_empty_row(sheet)

    try:
        for row in range(1, other_sheet.max_row + 1):
            for column in range(1, 5):
                sheet.cell(row_other, column).value = other_sheet.cell(row, column).value

            row_other += 1

    finally:
        other_book.save("other.xlsx")
        other_book.close()


@open_and_closing_data
def to_json(sheet):
    file = open("result.json", "w")
    data = {}
    try:
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row, 1).value:
                name = sheet.cell(row, 1).value
                birth_date = sheet.cell(row, 3).value
                gender = sheet.cell(row, 2).value
                death_date = sheet.cell(row, 4).value

                if death_date is not None:
                    death_date = function.date(death_date)

                data[row] = {
                    "Name": name,
                    "gender": gender,
                    "birth": function.date(birth_date),
                    "Dead": death_date
                }
        print(data)
        json.dump(data, file)

    finally:
        file.close()


if __name__ == "__main__":
    main()
